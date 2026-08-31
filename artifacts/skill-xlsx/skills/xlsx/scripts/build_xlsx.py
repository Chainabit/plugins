#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Build an .xlsx workbook from a JSON spec, using openpyxl.

The point of the spec is the column *type*. A spreadsheet whose numbers are stored
as text looks correct on screen and is useless the moment anyone sorts, sums, or
charts it — and that is the defect that survives review, because nothing about it
is visible. So every column declares a type, values are coerced to it, and a value
that cannot be coerced is a reported error rather than a silent string.

The whole spec is validated before the workbook is created, and every problem is
printed at once as `ERROR: <field>: <reason>`.

Usage:
    python3 build_xlsx.py spec.json output.xlsx [--validate-only]

Spec shape (see SKILL.md for the annotated version):

    {
      "properties": {"title": "...", "creator": "..."},   optional
      "sheets": [                                          required, at least one
        {
          "name": "Sales",                                 required
          "columns": [                                     required, at least one
            {"header": "Region", "key": "region", "type": "text", "width": 22},
            {"header": "Volume", "key": "volume", "type": "number",
             "format": "#,##0"}
          ],
          "rows": [                                        required
            {"region": "Istanbul", "volume": 4210},        objects keyed by "key"
            ["Izmir", 1904]                                or positional arrays
          ],
          "freezeHeader": true,                            optional, default true
          "autoFilter": true                               optional, default true
        }
      ]
    }
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys

COLUMN_TYPES = ("text", "number", "integer", "date", "datetime", "boolean", "formula")

# Default number formats per type, applied when a column declares no "format".
DEFAULT_FORMATS = {
    "number": "General",
    "integer": "0",
    "date": "yyyy-mm-dd",
    "datetime": "yyyy-mm-dd hh:mm",
}

# Excel refuses these in a sheet name, and silently truncates past 31 characters.
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

MAX_AUTO_WIDTH = 60
MIN_AUTO_WIDTH = 9
DEFAULT_FONT = os.environ.get("CHAINABIT_ARTIFACT_FONT_FAMILY", "IBM Plex Sans").strip() or "IBM Plex Sans"
ARABIC_FALLBACK_FONT = "IBM Plex Sans Arabic"
ARABIC_TEXT = re.compile(r"[\u0600-\u06ff\u0750-\u077f\u08a0-\u08ff]")
SAFE_FONT_NAME = re.compile(r"^[^\x00-\x1f\x7f]{1,80}$")


# --- validation -------------------------------------------------------------------


def validate_spec(spec: object) -> list[str]:
    """Return one `<field>: <reason>` string per problem; empty means the spec is usable."""
    problems: list[str] = []

    if not isinstance(spec, dict):
        return ["spec: top level must be a JSON object"]

    if spec.get("font") is not None and (
        not isinstance(spec.get("font"), str)
        or not SAFE_FONT_NAME.fullmatch(spec["font"].strip())
    ):
        problems.append("font: must be a safe non-empty family name")

    properties = spec.get("properties")
    if properties is not None:
        if not isinstance(properties, dict):
            problems.append("properties: must be an object when present")
        else:
            for key, value in properties.items():
                if key not in ("title", "creator", "subject", "description"):
                    problems.append(
                        f"properties.{key}: unknown property "
                        "(allowed: title, creator, subject, description)"
                    )
                elif not isinstance(value, str):
                    problems.append(f"properties.{key}: must be a string")

    sheets = spec.get("sheets")
    if not isinstance(sheets, list):
        problems.append("sheets: required, must be an array")
        return problems
    if not sheets:
        problems.append("sheets: must contain at least one sheet")
        return problems

    seen_names: set[str] = set()
    for index, sheet in enumerate(sheets):
        problems.extend(validate_sheet(sheet, f"sheets[{index}]", seen_names))

    return problems


def validate_sheet(sheet: object, where: str, seen_names: set[str]) -> list[str]:
    problems: list[str] = []

    if not isinstance(sheet, dict):
        return [f"{where}: must be an object"]

    name = sheet.get("name")
    if not isinstance(name, str) or not name.strip():
        problems.append(f"{where}.name: required, must be a non-empty string")
    else:
        if len(name) > 31:
            problems.append(f"{where}.name: must be at most 31 characters, found {len(name)}")
        if INVALID_SHEET_CHARS.search(name):
            problems.append(f'{where}.name: must not contain any of \\ / * ? : [ ]')
        lowered = name.strip().lower()
        if lowered in seen_names:
            problems.append(f"{where}.name: duplicate sheet name {name!r}")
        seen_names.add(lowered)

    for flag in ("freezeHeader", "autoFilter"):
        value = sheet.get(flag, True)
        if not isinstance(value, bool):
            problems.append(f"{where}.{flag}: must be true or false")

    columns = sheet.get("columns")
    keys: list[str | None] = []
    if not isinstance(columns, list) or not columns:
        problems.append(f"{where}.columns: required, must be a non-empty array")
        columns = []
    else:
        seen_keys: set[str] = set()
        for i, column in enumerate(columns):
            column_where = f"{where}.columns[{i}]"
            if not isinstance(column, dict):
                problems.append(f"{column_where}: must be an object")
                keys.append(None)
                continue

            header = column.get("header")
            if not isinstance(header, str) or not header.strip():
                problems.append(f"{column_where}.header: required, must be a non-empty string")

            kind = column.get("type", "text")
            if kind not in COLUMN_TYPES:
                problems.append(
                    f"{column_where}.type: must be one of {', '.join(COLUMN_TYPES)}, "
                    f"found {kind!r}"
                )

            key = column.get("key")
            if key is not None and not isinstance(key, str):
                problems.append(f"{column_where}.key: must be a string when present")
                key = None
            elif isinstance(key, str):
                if key in seen_keys:
                    problems.append(f"{column_where}.key: duplicate key {key!r}")
                seen_keys.add(key)
            keys.append(key if isinstance(key, str) else None)

            width = column.get("width")
            if width is not None:
                if not isinstance(width, (int, float)) or isinstance(width, bool):
                    problems.append(f"{column_where}.width: must be a number of characters")
                elif not 1 <= width <= 255:
                    problems.append(f"{column_where}.width: must be between 1 and 255")

            fmt = column.get("format")
            if fmt is not None and not isinstance(fmt, str):
                problems.append(f"{column_where}.format: must be an Excel format string")

    rows = sheet.get("rows")
    if not isinstance(rows, list):
        problems.append(f"{where}.rows: required, must be an array")
        return problems

    for i, row in enumerate(rows):
        problems.extend(validate_row(row, columns, keys, f"{where}.rows[{i}]"))

    return problems


def validate_row(row: object, columns: list, keys: list, where: str) -> list[str]:
    problems: list[str] = []

    if isinstance(row, list):
        if len(row) != len(columns):
            problems.append(
                f"{where}: has {len(row)} cells but {len(columns)} columns are declared"
            )
            return problems
        values = row
    elif isinstance(row, dict):
        missing = [k for k in keys if k is None]
        if missing:
            problems.append(
                f"{where}: given as an object, but not every column declares a "
                '"key" to look up. Add "key" to each column, or give rows as arrays.'
            )
            return problems
        values = [row.get(key) for key in keys]
        unknown = set(row) - {k for k in keys if k}
        if unknown:
            problems.append(
                f"{where}: has key(s) {', '.join(sorted(unknown))} matching no column"
            )
    else:
        return [f"{where}: must be an array of cells or an object keyed by column key"]

    for i, value in enumerate(values):
        if not isinstance(columns[i], dict):
            continue
        kind = columns[i].get("type", "text")
        reason = coercion_error(value, kind)
        if reason:
            problems.append(f"{where}[{i}]: {reason}")

    return problems


def coercion_error(value: object, kind: str) -> str | None:
    """Return why a value cannot be stored as `kind`, or None when it can."""
    if value is None:
        return None  # An empty cell is legitimate for every type.

    if kind == "text":
        if isinstance(value, (dict, list)):
            return "must be a scalar for a text column"
        return None

    if kind in ("number", "integer"):
        if isinstance(value, bool) or not isinstance(value, (int, float, str)):
            return f"must be a number for a {kind} column, found {type(value).__name__}"
        if isinstance(value, str):
            try:
                parsed = float(value.replace(",", "").strip())
            except ValueError:
                return (
                    f"{value!r} is not a number. A number column stores numbers; "
                    "use a text column if the value is genuinely a label."
                )
            if kind == "integer" and parsed != int(parsed):
                return f"{value!r} is not a whole number for an integer column"
        elif kind == "integer" and float(value) != int(value):
            return f"{value!r} is not a whole number for an integer column"
        return None

    if kind in ("date", "datetime"):
        if not isinstance(value, str):
            return f"must be an ISO 8601 string for a {kind} column, e.g. \"2026-08-18\""
        if parse_temporal(value, kind) is None:
            return (
                f"{value!r} is not ISO 8601. Use \"YYYY-MM-DD\" for a date or "
                '"YYYY-MM-DDTHH:MM:SS" for a datetime.'
            )
        return None

    if kind == "boolean":
        if not isinstance(value, bool):
            return f"must be true or false for a boolean column, found {value!r}"
        return None

    if kind == "formula":
        if not isinstance(value, str) or not value.startswith("="):
            return 'must be a string starting with "=" for a formula column'
        return None

    return None


def parse_temporal(value: str, kind: str):
    text = value.strip().replace("Z", "+00:00")
    try:
        if kind == "date":
            return dt.date.fromisoformat(text[:10])
        parsed = dt.datetime.fromisoformat(text)
        # Excel has no timezone concept; drop it rather than write a value the
        # reader will interpret in an unrelated zone.
        return parsed.replace(tzinfo=None)
    except ValueError:
        return None


def coerce(value: object, kind: str):
    if value is None:
        return None
    if kind == "text":
        return str(value)
    if kind == "number":
        return float(str(value).replace(",", "").strip()) if isinstance(value, str) else value
    if kind == "integer":
        return int(float(str(value).replace(",", "").strip())) if isinstance(value, str) else int(value)
    if kind in ("date", "datetime"):
        return parse_temporal(value, kind)
    return value


# --- rendering --------------------------------------------------------------------


def row_values(row, columns: list, keys: list) -> list:
    if isinstance(row, list):
        return list(row)
    return [row.get(key) for key in keys]


def auto_width(header: str, values: list) -> float:
    widest = len(header)
    for value in values:
        if value is None:
            continue
        widest = max(widest, len(str(value)))
    # +3 leaves room for the autofilter arrow, which otherwise sits on top of the
    # last character of the header.
    return max(MIN_AUTO_WIDTH, min(MAX_AUTO_WIDTH, widest + 3))


def build_workbook(spec: dict, output: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    # Workbook() starts with one sheet; the first spec sheet takes it over so the
    # file never ships with a stray empty "Sheet".
    workbook.remove(workbook.active)
    font_family = spec.get("font") or DEFAULT_FONT
    workbook._named_styles["Normal"].font = Font(name=font_family)

    properties = spec.get("properties") or {}
    if properties.get("title"):
        workbook.properties.title = properties["title"]
    if properties.get("creator"):
        workbook.properties.creator = properties["creator"]
    if properties.get("subject"):
        workbook.properties.subject = properties["subject"]
    if properties.get("description"):
        workbook.properties.description = properties["description"]

    header_fonts = {
        False: Font(name=font_family, bold=True, color="FFFFFF"),
        True: Font(name=ARABIC_FALLBACK_FONT, bold=True, color="FFFFFF"),
    }
    body_fonts = {
        False: Font(name=font_family),
        True: Font(name=ARABIC_FALLBACK_FONT),
    }
    header_fill = PatternFill("solid", fgColor="374151")
    header_alignment = Alignment(vertical="center", wrap_text=True)

    for sheet_spec in spec["sheets"]:
        sheet = workbook.create_sheet(title=sheet_spec["name"])
        columns = sheet_spec["columns"]
        keys = [column.get("key") for column in columns]
        rows = sheet_spec.get("rows", [])

        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=index, value=column["header"])
            cell.font = header_fonts[bool(ARABIC_TEXT.search(str(column["header"])))]
            cell.fill = header_fill
            cell.alignment = header_alignment
        sheet.row_dimensions[1].height = 20

        for row_index, row in enumerate(rows, start=2):
            values = row_values(row, columns, keys)
            for column_index, (column, value) in enumerate(zip(columns, values), start=1):
                kind = column.get("type", "text")
                cell = sheet.cell(row=row_index, column=column_index,
                                  value=coerce(value, kind))
                cell.font = body_fonts[bool(ARABIC_TEXT.search(str(value or "")))]
                number_format = column.get("format") or DEFAULT_FORMATS.get(kind)
                if number_format:
                    cell.number_format = number_format

        for index, column in enumerate(columns, start=1):
            letter = get_column_letter(index)
            width = column.get("width")
            if width is None:
                key = column.get("key")
                sample = [
                    row_values(row, columns, keys)[index - 1]
                    for row in rows[:200]
                ]
                width = auto_width(column["header"], sample)
            sheet.column_dimensions[letter].width = width

        last_column = get_column_letter(len(columns))
        if sheet_spec.get("freezeHeader", True):
            sheet.freeze_panes = "A2"
        if sheet_spec.get("autoFilter", True) and rows:
            sheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"

    workbook.save(output)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="build_xlsx.py",
        description=(
            "Build an .xlsx workbook from a JSON spec: typed cells, a styled header "
            "row, column widths, and per-column number formats."
        ),
        epilog=(
            "Example:\n"
            "  python3 build_xlsx.py spec.json sales.xlsx\n"
            "  python3 build_xlsx.py spec.json sales.xlsx --validate-only\n\n"
            "The spec is fully validated first. Every problem is printed as an "
            "ERROR: line so the spec can be corrected in one pass.\n"
            "Runs offline. openpyxl is already installed."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("spec", help="path to the JSON spec file")
    parser.add_argument("output", help="path to write the .xlsx to")
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="check the spec and exit without writing a workbook",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    problems: list[str] = []

    if not args.output.lower().endswith(".xlsx"):
        problems.append(
            f"output: {args.output} must end in .xlsx — this script does not write "
            "CSV, and renaming an .xlsx does not convert it"
        )
    output_dir = os.path.dirname(os.path.abspath(args.output))
    if not os.path.isdir(output_dir):
        problems.append(f"output: directory {output_dir} does not exist")
    elif not os.access(output_dir, os.W_OK):
        problems.append(f"output: directory {output_dir} is not writable")

    try:
        with open(args.spec, "r", encoding="utf-8") as handle:
            spec = json.load(handle)
    except FileNotFoundError:
        print(f"ERROR: spec: {args.spec} does not exist", file=sys.stderr)
        return 1
    except IsADirectoryError:
        print(f"ERROR: spec: {args.spec} is a directory, expected a JSON file", file=sys.stderr)
        return 1
    except PermissionError:
        print(f"ERROR: spec: no permission to read {args.spec}", file=sys.stderr)
        return 1
    except UnicodeDecodeError as exc:
        print(f"ERROR: spec: {args.spec} is not valid UTF-8 ({exc.reason})", file=sys.stderr)
        return 1
    except json.JSONDecodeError as exc:
        print(
            f"ERROR: spec: {args.spec} is not valid JSON — {exc.msg} "
            f"at line {exc.lineno}, column {exc.colno}",
            file=sys.stderr,
        )
        return 1

    problems.extend(validate_spec(spec))

    if problems:
        for problem in problems:
            print(f"ERROR: {problem}", file=sys.stderr)
        return 1

    if args.validate_only:
        total_rows = sum(len(sheet.get("rows", [])) for sheet in spec["sheets"])
        print(
            f"OK: {args.spec} is a valid workbook spec "
            f"({len(spec['sheets'])} sheet(s), {total_rows} row(s))"
        )
        return 0

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print(
            "ERROR: environment: the 'openpyxl' package is missing. This script only "
            "runs inside the Chainabit sandbox image, where it is pre-installed. Do "
            "not try to install it.",
            file=sys.stderr,
        )
        return 2

    try:
        build_workbook(spec, args.output)
    except PermissionError:
        print(f"ERROR: output: no permission to write {args.output}", file=sys.stderr)
        return 2
    except OSError as exc:
        print(f"ERROR: output: could not write {args.output}: {exc}", file=sys.stderr)
        return 2

    size = os.path.getsize(args.output)
    print(f"OK: wrote {args.output} ({size} bytes)")
    print(f"Next: python3 scripts/validate_xlsx.py {args.output}")
    with open(args.output, "rb") as handle:
        digest = hashlib.sha256(handle.read()).hexdigest()
    font = spec.get("font") or DEFAULT_FONT
    print(json.dumps({
        "schema": "chainabit.xlsx.execution/v1",
        "success": True,
        "generator": "skill-xlsx.build_xlsx",
        "output": {
            "path": os.path.realpath(args.output),
            "shape": "file",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "sha256": digest,
            "bytes": size,
        },
        "typography": {
            "family": font,
            "source": "user_override" if spec.get("font") else "chainabit_default",
        },
    }, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
